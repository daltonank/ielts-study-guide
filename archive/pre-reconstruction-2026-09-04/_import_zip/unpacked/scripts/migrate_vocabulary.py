#!/usr/bin/env python3
"""
Deterministic Phase 2 vocabulary migration.
Usage:
  python scripts/migrate_vocabulary.py path/to/IELTS_Academic_C1_Ukrainian_Vocabulary_Bank.xlsx

The script refuses to mark migration complete unless the normalized Study Bank has exactly 1,784 unique records.
"""
from pathlib import Path
import sys, json, re
from openpyxl import load_workbook

EXPECTED = 1784

def norm(x):
    if x is None:
        return ""
    return re.sub(r"\s+", " ", str(x).strip())

def clean_value(x):
    x = norm(x)
    if x.startswith('='):
        return None
    return x or None

def rows(ws):
    return list(ws.iter_rows(values_only=True))

def source_maps(wb):
    ox_map, awl_map, starter = {}, {}, {}
    # Oxford
    for row in rows(wb['Oxford C1 Bank'])[1:]:
        if not row or not row[1]:
            continue
        key = norm(row[1]).casefold()
        ox_map[key] = {
            'sourceSheet':'Oxford C1 Bank','sourceId':row[0],'word':norm(row[1]),'pos':norm(row[2]),'ua':norm(row[3]),
            'awlOverlap':norm(row[4]),'awlSublist':row[5],'priority':norm(row[6]),'topic':norm(row[7]),
            'studyStatus':norm(row[8]),'confidence':row[9] if row[9] is not None else 0,
            'lastReviewed':clean_value(row[10]),'nextReview':clean_value(row[11]),'ownExample':clean_value(row[12]),
            'collocationNote':clean_value(row[13]),'translationQa':clean_value(row[14]),'sourceUrl':clean_value(row[15])
        }
    # AWL
    for row in rows(wb['Academic Word List'])[1:]:
        if not row or not row[1]:
            continue
        key = norm(row[1]).casefold()
        awl_map[key] = {
            'sourceSheet':'Academic Word List','sourceId':row[0],'word':norm(row[1]),'ua':norm(row[2]),'awlSublist':row[3],
            'oxfordOverlap':norm(row[4]),'priority':norm(row[5]),'topic':norm(row[6]),'studyStatus':norm(row[7]),
            'confidence':row[8] if row[8] is not None else 0,'lastReviewed':clean_value(row[9]),'nextReview':clean_value(row[10]),
            'ownExample':clean_value(row[11]),'collocationNote':clean_value(row[12]),'translationQa':clean_value(row[13]),'sourceUrl':clean_value(row[14])
        }
    # Starter 100
    for row in rows(wb['Starter 100'])[4:]:
        if not row or not row[2]:
            continue
        starter[norm(row[2]).casefold()] = {
            'starterSequence': row[0], 'starterSource': norm(row[1]), 'starterTopic': norm(row[5]),
            'starterChecked': norm(row[9])
        }
    return ox_map, awl_map, starter

def parse_source_parts(label):
    label = norm(label)
    if not label:
        return []
    return [x.strip() for x in label.split('+')]

def dedupe_keep(seq):
    out=[]
    seen=set()
    for item in seq:
        if item is None or item=="":
            continue
        if item not in seen:
            seen.add(item); out.append(item)
    return out

def map_legacy_status(status):
    s = norm(status).lower()
    if s in {'', 'not started'}:
        return 'New'
    if s in {'learning', 'in progress'}:
        return 'Recognized'
    if s in {'reviewed'}:
        return 'Recall'
    if s in {'mastered'}:
        return 'Active'
    return 'New'

def split_collocations(text):
    if not text:
        return []
    parts = re.split(r';|,\s*(?=[A-Za-z])', text)
    return [p.strip() for p in parts if p and p.strip()]

def main(path):
    path = Path(path)
    if not path.exists():
        raise SystemExit(f'Missing workbook: {path}')
    wb = load_workbook(path, read_only=True, data_only=False)
    ox_map, awl_map, starter_map = source_maps(wb)

    study_rows = rows(wb['Study Bank'])
    headers = study_rows[3]
    if norm(headers[0]).lower() != 'english word':
        raise SystemExit('G2 FAIL: Study Bank header row not found at row 4.')

    records = []
    seen = set()
    for idx, row in enumerate(study_rows[4:], start=1):
        if not row or not row[0]:
            continue
        word = norm(row[0])
        key = word.casefold()
        if key in seen:
            raise SystemExit(f'G2 FAIL: duplicate normalized headword in Study Bank: {word}')
        seen.add(key)
        pos = norm(row[1])
        ua = norm(row[2])
        definitionUa = norm(row[3])
        sourceLabel = norm(row[4])
        priority = norm(row[5])
        legacyStatus = norm(row[6])
        confidence = row[7] if len(row) > 7 and row[7] is not None else 0

        ox = ox_map.get(key)
        awl = awl_map.get(key)
        starter = starter_map.get(key)

        topics = dedupe_keep([ox.get('topic') if ox else None, awl.get('topic') if awl else None, starter.get('starterTopic') if starter else None])
        source_urls = dedupe_keep([ox.get('sourceUrl') if ox else None, awl.get('sourceUrl') if awl else None])
        review_candidates = [ox.get('lastReviewed') if ox else None, awl.get('lastReviewed') if awl else None]
        next_candidates = [ox.get('nextReview') if ox else None, awl.get('nextReview') if awl else None]
        examples = dedupe_keep([ox.get('ownExample') if ox else None, awl.get('ownExample') if awl else None])
        colloc_text = dedupe_keep([ox.get('collocationNote') if ox else None, awl.get('collocationNote') if awl else None])
        translation_qas = dedupe_keep([ox.get('translationQa') if ox else None, awl.get('translationQa') if awl else None])

        rec = {
            'id': f'SB-{idx:04d}',
            'legacyRow': idx + 4,
            'word': word,
            'pos': pos or (ox.get('pos') if ox else '') or (awl.get('pos') if awl else ''),
            'ua': ua,
            'definitionUa': definitionUa,
            'source': sourceLabel,
            'sourceParts': parse_source_parts(sourceLabel),
            'priority': priority,
            'topic': topics[0] if len(topics)==1 else (' / '.join(topics) if topics else ''),
            'topicTags': topics,
            'legacyStatus': legacyStatus,
            'confidence': confidence,
            'lastReviewed': next((x for x in review_candidates if x), None),
            'nextReview': next((x for x in next_candidates if x), None),
            'example': examples[0] if examples else '',
            'collocationNote': ' | '.join(colloc_text),
            'collocations': split_collocations(colloc_text[0]) if colloc_text else [],
            'register': 'academic/neutral',
            'masteryDefault': map_legacy_status(legacyStatus),
            'starter100': bool(starter),
            'starterSequence': starter.get('starterSequence') if starter else None,
            'starterSource': starter.get('starterSource') if starter else None,
            'sourceRefs': {
                'oxfordId': ox.get('sourceId') if ox else None,
                'awlId': awl.get('sourceId') if awl else None,
                'awlSublist': awl.get('awlSublist') if awl else (ox.get('awlSublist') if ox else None),
                'oxfordOverlap': awl.get('oxfordOverlap') if awl else None,
                'awlOverlap': ox.get('awlOverlap') if ox else None,
            },
            'translationQa': ' | '.join(translation_qas),
            'sourceUrls': source_urls,
            'originality': 'legacy-workbook'
        }
        if not rec['ua']:
            raise SystemExit(f"G2 FAIL: blank Ukrainian equivalent for {word}")
        records.append(rec)

    if len(records) != EXPECTED:
        raise SystemExit(f'G2 FAIL: expected {EXPECTED} normalized records, found {len(records)}')

    # Additional reconciliation checks against source totals.
    ox_count = sum(1 for r in records if 'Oxford C1' in r['source'])
    awl_count = sum(1 for r in records if 'AWL' in r['source'])
    starter_count = sum(1 for r in records if r['starter100'])
    meta = {
        'complete': True,
        'expectedCount': EXPECTED,
        'sourceCount': len(records),
        'sourceWorkbook': path.name,
        'sourceBreakdown': {'Oxford-tagged': ox_count, 'AWL-tagged': awl_count, 'Starter100': starter_count},
        'gate': 'G2 SOURCE RECONCILED'
    }

    out = Path(__file__).resolve().parents[1] / 'web' / 'vocabulary.js'
    out.write_text("window.VOCABULARY_META=" + json.dumps(meta, ensure_ascii=False) + ";\n" +
                   "window.VOCABULARY=" + json.dumps(records, ensure_ascii=False) + ";\n",
                   encoding='utf-8')

    manifest = Path(__file__).resolve().parents[1] / 'docs' / 'vocabulary_migration_manifest.json'
    manifest.write_text(json.dumps({'meta': meta, 'ids': [r['id'] for r in records], 'sample': records[:5]}, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f"G2 PASS: {len(records)}/{EXPECTED} records migrated -> {out}")

if __name__ == '__main__':
    if len(sys.argv) != 2:
        raise SystemExit(__doc__)
    main(sys.argv[1])
